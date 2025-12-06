"""
Helper para exportar reportes a PDF y Excel
"""
from io import BytesIO
from typing import List, Dict, Any, Optional
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ExportHelper:
    """Clase helper para exportar reportes a diferentes formatos"""
    
    @staticmethod
    def export_to_excel(
        data: List[Dict[str, Any]],
        title: str,
        filename: str = "reporte.xlsx",
        headers: Optional[List[str]] = None
    ) -> BytesIO:
        """
        Exporta datos a un archivo Excel
        
        Args:
            data: Lista de diccionarios con los datos
            title: Título del reporte
            filename: Nombre del archivo
            headers: Lista de encabezados (opcional, se infiere de los datos)
        
        Returns:
            BytesIO con el contenido del archivo Excel
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        title_font = Font(bold=True, size=16)
        title_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws.merge_cells('A1:' + get_column_letter(len(data[0]) if data else 5) + '1')
        title_cell = ws['A1']
        title_cell.value = title
        title_cell.font = title_font
        title_cell.alignment = title_alignment
        
        # Fecha de generación
        ws.merge_cells('A2:' + get_column_letter(len(data[0]) if data else 5) + '2')
        date_cell = ws['A2']
        date_cell.value = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        date_cell.alignment = title_alignment
        
        if not data:
            return ExportHelper._workbook_to_bytes(wb)
        
        # Encabezados
        if not headers:
            headers = list(data[0].keys())
        
        row_num = 4
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = header.upper().replace('_', ' ')
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Datos
        for row_idx, row_data in enumerate(data, row_num + 1):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = row_data.get(header, '')
                
                # Formatear valores
                if isinstance(value, (int, float)):
                    cell.value = value
                    cell.number_format = '#,##0.00' if isinstance(value, float) else '#,##0'
                else:
                    cell.value = str(value) if value is not None else ''
                
                cell.border = border
                cell.alignment = Alignment(vertical="center")
        
        # Ajustar ancho de columnas
        for col_idx in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for cell in ws[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        return ExportHelper._workbook_to_bytes(wb)
    
    @staticmethod
    def _workbook_to_bytes(workbook: Workbook) -> BytesIO:
        """Convierte un workbook a BytesIO"""
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def export_to_pdf(
        data: List[Dict[str, Any]],
        title: str,
        headers: Optional[List[str]] = None,
        page_size=A4
    ) -> BytesIO:
        """
        Exporta datos a un archivo PDF
        
        Args:
            data: Lista de diccionarios con los datos
            title: Título del reporte
            headers: Lista de encabezados (opcional, se infiere de los datos)
            page_size: Tamaño de página (default: A4)
        
        Returns:
            BytesIO con el contenido del archivo PDF
        """
        buffer = BytesIO()
        # Aumentar márgenes para mejor presentación
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=page_size,
            leftMargin=50,
            rightMargin=50,
            topMargin=50,
            bottomMargin=50
        )
        elements = []
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#667EEA'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Título
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 0.1 * inch))
        elements.append(Paragraph(
            f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            styles['Normal']
        ))
        elements.append(Spacer(1, 0.4 * inch))
        
        if not data:
            elements.append(Paragraph("No hay datos para mostrar", styles['Normal']))
            doc.build(elements)
            buffer.seek(0)
            return buffer
        
        # Preparar datos para la tabla
        if not headers:
            headers = list(data[0].keys())
        
        # Encabezados formateados
        table_data = [[h.upper().replace('_', ' ') for h in headers]]
        
        # Datos
        for row in data:
            table_row = []
            for header in headers:
                value = row.get(header, '')
                if isinstance(value, float):
                    table_row.append(f"{value:,.2f}")
                elif isinstance(value, int):
                    table_row.append(f"{value:,}")
                else:
                    table_row.append(str(value) if value is not None else '')
            table_data.append(table_row)
        
        # Crear tabla con anchos autoajustables
        # Calcular ancho disponible considerando márgenes
        available_width = page_size[0] - 100  # Restamos márgenes
        col_widths = [available_width / len(headers) for _ in headers]
        
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        
        # Estilo de tabla optimizado para evitar sobreposición
        table.setStyle(TableStyle([
            # Encabezados
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667EEA')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('LEFTPADDING', (0, 0), (-1, 0), 5),
            ('RIGHTPADDING', (0, 0), (-1, 0), 5),
            
            # Datos
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 6),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('LEFTPADDING', (0, 1), (-1, -1), 5),
            ('RIGHTPADDING', (0, 1), (-1, -1), 5),
            ('WORDWRAP', (0, 0), (-1, -1), True),
        ]))
        
        elements.append(table)
        
        # Construir PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def export_attendance_report_pdf(
        attendance_data: List[Dict[str, Any]],
        start_date: str,
        end_date: str
    ) -> BytesIO:
        """Exporta reporte de asistencias a PDF con anchos personalizados"""
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.enums import TA_CENTER
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=A4,
            leftMargin=40,
            rightMargin=40,
            topMargin=50,
            bottomMargin=50
        )
        elements = []
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#667EEA'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Título
        title = f"Reporte de Asistencias {start_date} al {end_date}"
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 0.1 * inch))
        elements.append(Paragraph(
            f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            styles['Normal']
        ))
        elements.append(Spacer(1, 0.4 * inch))
        
        if not attendance_data:
            elements.append(Paragraph("No hay datos para mostrar", styles['Normal']))
            doc.build(elements)
            buffer.seek(0)
            return buffer
        
        # Encabezados más cortos
        headers_display = ['EMPLEADO', 'DEPARTAMENTO', 'PRESENTE', 'AUSENTE', 'TARDAN.', '% ASIST.']
        headers_data = ['empleado', 'departamento', 'presente', 'ausente', 'tardanzas', 'porcentaje_asistencia']
        
        # Preparar datos
        table_data = [headers_display]
        
        for row in attendance_data:
            table_row = [
                str(row.get('empleado', '')),
                str(row.get('departamento', '')),
                str(row.get('presente', 0)),
                str(row.get('ausente', 0)),
                str(row.get('tardanzas', 0)),
                f"{row.get('porcentaje_asistencia', 0):.1f}"
            ]
            table_data.append(table_row)
        
        # Anchos autoajustables - más espacio para nombres largos
        col_widths = [160, 100, 50, 50, 50, 50]  # Total: 460
        
        table = Table(table_data, colWidths=col_widths)
        
        # Estilo de tabla con más espacio
        table.setStyle(TableStyle([
            # Encabezados
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667EEA')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (1, -1), 'LEFT'),  # Empleado y Departamento alineados a la izquierda
            ('ALIGN', (2, 0), (-1, -1), 'CENTER'),  # Números centrados
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('LEFTPADDING', (0, 0), (-1, 0), 6),
            ('RIGHTPADDING', (0, 0), (-1, 0), 6),
            
            # Datos
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('LEFTPADDING', (0, 1), (-1, -1), 6),
            ('RIGHTPADDING', (0, 1), (-1, -1), 6),
            ('WORDWRAP', (0, 0), (-1, -1), True),  # Permitir salto de línea
        ]))
        
        elements.append(table)
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def export_attendance_report_excel(
        attendance_data: List[Dict[str, Any]],
        start_date: str,
        end_date: str
    ) -> BytesIO:
        """Exporta reporte de asistencias a Excel con formato mejorado"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Asistencias"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        title_font = Font(bold=True, size=16)
        title_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f"Reporte de Asistencias - {start_date} al {end_date}"
        title_cell.font = title_font
        title_cell.alignment = title_alignment
        
        # Fecha
        ws.merge_cells('A2:F2')
        date_cell = ws['A2']
        date_cell.value = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        date_cell.alignment = title_alignment
        
        # Encabezados
        headers = ['EMPLEADO', 'DEPARTAMENTO', 'PRESENTE', 'AUSENTE', 'TARDANZAS', '% ASISTENCIA']
        row_num = 4
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Datos
        for row_idx, row_data in enumerate(attendance_data, row_num + 1):
            ws.cell(row=row_idx, column=1, value=row_data.get('empleado', ''))
            ws.cell(row=row_idx, column=2, value=row_data.get('departamento', ''))
            ws.cell(row=row_idx, column=3, value=row_data.get('presente', 0))
            ws.cell(row=row_idx, column=4, value=row_data.get('ausente', 0))
            ws.cell(row=row_idx, column=5, value=row_data.get('tardanzas', 0))
            ws.cell(row=row_idx, column=6, value=row_data.get('porcentaje_asistencia', 0))
            
            # Aplicar bordes y formato
            for col_num in range(1, 7):
                cell = ws.cell(row=row_idx, column=col_num)
                cell.border = border
                cell.alignment = Alignment(horizontal="center" if col_num > 2 else "left", vertical="center")
                if col_num == 6:
                    cell.number_format = '0.0'
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 30  # Empleado más ancho
        ws.column_dimensions['B'].width = 25  # Departamento
        ws.column_dimensions['C'].width = 12  # Presente
        ws.column_dimensions['D'].width = 12  # Ausente
        ws.column_dimensions['E'].width = 12  # Tardanzas
        ws.column_dimensions['F'].width = 15  # % Asistencia
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def export_employee_report_pdf(employees: List[Dict[str, Any]]) -> BytesIO:
        """Exporta reporte de empleados a PDF"""
        title = "Reporte General de Empleados"
        headers = ['id_empleado', 'nombre', 'apellido', 'email', 'departamento', 'puesto', 'salario', 'fecha_ingreso']
        return ExportHelper.export_to_pdf(employees, title, headers)
    
    @staticmethod
    def export_employee_report_excel(employees: List[Dict[str, Any]]) -> BytesIO:
        """Exporta reporte de empleados a Excel"""
        title = "Reporte General de Empleados"
        headers = ['id_empleado', 'nombre', 'apellido', 'email', 'departamento', 'puesto', 'salario', 'fecha_ingreso']
        return ExportHelper.export_to_excel(employees, title, "reporte_empleados.xlsx", headers)
    
    @staticmethod
    def export_payroll_report_pdf(
        payroll_data: List[Dict[str, Any]],
        month: int,
        year: int
    ) -> BytesIO:
        """Exporta reporte de nómina a PDF"""
        meses = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
        title = f"Reporte de Nómina - {meses[month]} {year}"
        headers = ['empleado', 'departamento', 'salario_base', 'bonificaciones', 'deducciones', 'total']
        return ExportHelper.export_to_pdf(payroll_data, title, headers)
    
    @staticmethod
    def export_payroll_report_excel(
        payroll_data: List[Dict[str, Any]],
        month: int,
        year: int
    ) -> BytesIO:
        """Exporta reporte de nómina a Excel"""
        meses = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
        title = f"Reporte de Nómina - {meses[month]} {year}"
        headers = ['empleado', 'departamento', 'salario_base', 'bonificaciones', 'deducciones', 'total']
        return ExportHelper.export_to_excel(payroll_data, title, "reporte_nomina.xlsx", headers)
    
    @staticmethod
    def export_vacation_report_pdf(vacations: List[Dict[str, Any]]) -> BytesIO:
        """Exporta reporte de vacaciones a PDF"""
        title = "Reporte de Vacaciones y Permisos"
        headers = ['empleado', 'tipo', 'fecha_inicio', 'fecha_fin', 'dias', 'estado', 'fecha_solicitud']
        return ExportHelper.export_to_pdf(vacations, title, headers)
    
    @staticmethod
    def export_vacation_report_excel(vacations: List[Dict[str, Any]]) -> BytesIO:
        """Exporta reporte de vacaciones a Excel"""
        title = "Reporte de Vacaciones y Permisos"
        headers = ['empleado', 'tipo', 'fecha_inicio', 'fecha_fin', 'dias', 'estado', 'fecha_solicitud']
        return ExportHelper.export_to_excel(vacations, title, "reporte_vacaciones.xlsx", headers)

